"""Модуль для работы с файлами Excel."""
import os
import time
from openpyxl import load_workbook, Workbook

class ExcelManager:
    """Управляет записью штрихкодов в файлы Excel."""
    
    def __init__(self, save_dir: str):
        self.save_dir = save_dir
        self.base_file = os.path.join(save_dir, "Mark_fail.xlsx")
        self.column = "A"

    def update_save_dir(self, new_dir: str) -> None:
        """Обновляет целевую директорию сохранения файлов."""
        self.save_dir = new_dir
        self.base_file = os.path.join(new_dir, "Mark_fail.xlsx")

    def init_workbook(self) -> Workbook:
        """Инициализирует или загружает временный файл."""
        if os.path.exists(self.base_file):
            try:
                return load_workbook(self.base_file)
            except Exception:
                return Workbook()
        return Workbook()

    def save_code(self, code: str, timestamp: str) -> None:
        """Записывает код маркировки и временную метку в Excel."""
        wb = self.init_workbook()
        ws = wb.active
        row = ws.max_row + 1
        
        ws[f'{self.column}{row}'] = code
        ws[f'B{row}'] = timestamp
        wb.save(self.base_file)

    def finalize_file(self, batch_name: str) -> str:
        """Переименовывает собранную коробку в указанную папку."""
        if not os.path.exists(self.base_file):
            return ""
            
        new_filename = os.path.join(self.save_dir, f"{batch_name}.xlsx")
        if os.path.exists(new_filename):
            timestamp = time.strftime('%Y%m%d_%H%M%S')
            new_filename = os.path.join(self.save_dir, f"{batch_name}_{timestamp}.xlsx")
        
        try:
            os.rename(self.base_file, new_filename)
            return new_filename
        except Exception:
            return ""
